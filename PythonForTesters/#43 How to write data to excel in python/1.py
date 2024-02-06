from openpyxl import Workbook
somedata = [[1,2],[3,4]]

wb = Workbook()
ws = wb.active
for data in somedata:
    ws.append(data)
for i in range(1,11):
    for u in range(1,11):
        ws.cell(row=i,column=u).value = i+u

ws['A11'] = "Data, written from python"
wb.save("fileTwo.xlsx")