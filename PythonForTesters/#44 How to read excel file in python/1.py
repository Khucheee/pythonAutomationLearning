from openpyxl import Workbook,load_workbook

wb = load_workbook(filename='fileTwo.xlsx')
ws = wb.active

print(ws['A1'].value)
print(wb['Лист1']['A2'].value)
print(ws.cell(1,3).value)
rows = ws.max_row
columns = ws.max_column
for i in range(1,rows+1):
    print()
    for j in range(1,columns+1):
        print(ws.cell(i,j).value, sep=' ', end=' ',)

