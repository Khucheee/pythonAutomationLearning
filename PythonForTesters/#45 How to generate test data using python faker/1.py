from faker import Faker
from openpyxl import Workbook
a = Faker(['ru'])
wb = Workbook()
ws = wb.active
for i in range(1,10):
    ws.cell(i,1).value = a.name()
    ws.cell(i,2).value = a.address()
wb.save('fileThree.xlsx')
